"""
Generic Excel Utility Library for SAP Data Analysis

This library provides reusable functions for creating professional Excel reports
with consistent formatting, suitable for SAP analysis and documentation tasks.

Author: Created with Claude Code
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from typing import List, Dict, Tuple


class ExcelReportBuilder:
    """Builder class for creating formatted Excel reports"""

    def __init__(self, title: str, subtitle: str = None):
        """
        Initialize a new Excel report

        Args:
            title: Main title for the report
            subtitle: Optional subtitle with metadata
        """
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.title = title
        self.subtitle = subtitle
        self._current_row = 1

    def add_title(self, merge_range: str = 'A1:I1'):
        """Add formatted title to the worksheet"""
        self.ws.merge_cells(merge_range)
        self.ws['A1'] = self.title
        self.ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='1F4E79')
        self._current_row = 2

    def add_subtitle(self, merge_range: str = 'A2:I2'):
        """Add formatted subtitle to the worksheet"""
        if self.subtitle:
            self.ws.merge_cells(merge_range)
            self.ws['A2'] = self.subtitle
            self.ws['A2'].font = Font(name='Calibri', size=10, italic=True, color='666666')
            self._current_row = 3

    def add_headers(self, headers: List[Tuple[str, str, int]], row: int = 4):
        """
        Add formatted column headers

        Args:
            headers: List of tuples (column_letter, header_text, width)
            row: Row number for headers (default: 4)
        """
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_align = Alignment(wrap_text=True, vertical='center', horizontal='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col_letter, header_text, width in headers:
            cell = self.ws[f'{col_letter}{row}']
            cell.value = header_text
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            self.ws.column_dimensions[col_letter].width = width

        self._current_row = row + 1

    def add_data_rows(self, data: List[Dict], column_mapping: Dict[str, str],
                     start_row: int = 5, apply_borders: bool = True):
        """
        Add data rows to the worksheet

        Args:
            data: List of dictionaries containing row data
            column_mapping: Dict mapping column letters to data keys
            start_row: Starting row number for data
            apply_borders: Whether to apply borders to cells
        """
        wrap_align = Alignment(wrap_text=True, vertical='top')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for idx, row_data in enumerate(data):
            row = start_row + idx
            for col_letter, data_key in column_mapping.items():
                cell = self.ws[f'{col_letter}{row}']
                cell.value = row_data.get(data_key, '')
                cell.alignment = wrap_align
                cell.font = Font(name='Calibri', size=10)
                if apply_borders:
                    cell.border = thin_border

        self._current_row = start_row + len(data)

    def apply_conditional_formatting(self, cell_ref: str, value: str,
                                    color_map: Dict[str, Tuple[str, str]]):
        """
        Apply conditional formatting based on cell value

        Args:
            cell_ref: Cell reference (e.g., 'F5')
            value: Value to check against color_map
            color_map: Dict mapping values to (bg_color, font_color) tuples
        """
        if value in color_map:
            bg_color, font_color = color_map[value]
            cell = self.ws[cell_ref]
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
            cell.font = Font(name='Calibri', size=10, bold=True, color=font_color)

    def set_row_height(self, row: int, height: int):
        """Set the height of a specific row"""
        self.ws.row_dimensions[row].height = height

    def freeze_panes(self, cell: str):
        """Freeze panes at the specified cell"""
        self.ws.freeze_panes = cell

    def add_auto_filter(self, range_ref: str):
        """Add auto-filter to the specified range"""
        self.ws.auto_filter.ref = range_ref

    def create_new_sheet(self, title: str):
        """Create a new worksheet and return it"""
        return self.wb.create_sheet(title)

    def save(self, filepath: str):
        """Save the workbook to the specified path"""
        self.wb.save(filepath)
        return filepath


class SeverityColorScheme:
    """Pre-defined color schemes for severity levels"""

    STANDARD = {
        'HIGH': ('FFC7CE', '9C0006'),      # Red background, dark red text
        'MEDIUM': ('FFEB9C', '9C6500'),    # Yellow background, dark yellow text
        'LOW': ('C6EFCE', '006100'),       # Green background, dark green text
    }

    @classmethod
    def get_fills(cls):
        """Get PatternFill objects for each severity level"""
        return {
            'HIGH': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
            'MEDIUM': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
            'LOW': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
        }


def generate_timestamp() -> str:
    """Generate a formatted timestamp for reports"""
    return datetime.now().strftime("%Y-%m-%d %H:%M")


def create_border(style: str = 'thin') -> Border:
    """Create a border with the specified style"""
    return Border(
        left=Side(style=style),
        right=Side(style=style),
        top=Side(style=style),
        bottom=Side(style=style)
    )
