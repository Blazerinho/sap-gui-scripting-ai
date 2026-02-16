import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Error Analysis'

# Styles
header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
title_font = Font(name='Calibri', size=14, bold=True, color='1F4E79')
subtitle_font = Font(name='Calibri', size=10, italic=True, color='666666')
wrap_align = Alignment(wrap_text=True, vertical='top')
header_align = Alignment(wrap_text=True, vertical='center', horizontal='center')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
severity_high_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
severity_med_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
severity_low_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

# Title
ws.merge_cells('A1:I1')
ws['A1'] = 'Posting Engine Error Monitor Analysis - Project ZFI_SDT'
ws['A1'].font = title_font

ws.merge_cells('A2:I2')
ws['A2'] = f'System: MWA / Client: 100 / Area: ZAP_OI (Vendor Open Items) / Scenario: Receiver Processing / Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
ws['A2'].font = subtitle_font

# Headers (row 4)
headers = [
    ('A', 'No.', 5),
    ('B', 'Message Class', 15),
    ('C', 'Message Number', 15),
    ('D', 'Count (RCV)', 13),
    ('E', 'Error Message', 45),
    ('F', 'Severity', 12),
    ('G', 'Root Cause Analysis', 50),
    ('H', 'Proposed Solution', 55),
    ('I', 'Reference / SAP Note', 25),
]

for col_letter, header_text, width in headers:
    cell = ws[f'{col_letter}4']
    cell.value = header_text
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border
    ws.column_dimensions[col_letter].width = width

# Error data with analysis - sorted by count (severity)
errors = [
    {
        'msg_class': 'NR',
        'msg_no': '751',
        'count': '9999+',
        'text': 'Interval does not exist for object (Number range missing)',
        'severity': 'HIGH',
        'root_cause': 'The document type used in target posting refers to a number range interval that does not exist or is not maintained in the target system (MWA client 100). With 9999+ occurrences this is the most widespread error and likely blocks the majority of worklist items from being posted.',
        'solution': '1. Identify which number range objects are missing: check the error details for the specific object name (e.g., FBNR for FI documents).\n2. Go to the target system and maintain the number range in customizing (e.g., FBN1 for FI document numbers, or SNRO for general number ranges).\n3. Ensure both the number range NUMBER and the FROM/TO interval are maintained.\n4. After fixing, rebuild transfer list and re-simulate affected worklist items.',
        'reference': 'PECON FAQ #35\nSNRO / FBN1 in target system'
    },
    {
        'msg_class': 'CNV_PE',
        'msg_no': '451',
        'count': '2328',
        'text': 'Transformation rule ended with error (see long text)',
        'severity': 'HIGH',
        'root_cause': 'One or more transformation rules (TRules) failed during the Build Transfer List step. Common causes:\n- Currency settings not read yet (WAERS rule)\n- Company code transferred does not exist in target\n- TRule runtime not generated\n- Missing data in source system for the transformation.',
        'solution': '1. Check the long text of the error for each affected worklist item in the Worklist Monitor to identify which specific TRule failed.\n2. If currency-related (_PE_FI_WAERS): Perform "Read Customizing" at area level in CNV_PE_PROJ -> Expert tab, with option "overwrite".\n3. If TRule runtime not generated: Go to CNV_PE_PROJ -> Area -> Execution Rule -> Transfer Method -> Transformation, find the failing TRule, activate it. Or run report CNV_PE_SUPPORT_PROJ_GEN_TRULES.\n4. Rebuild transfer list after fixing.',
        'reference': 'PECON FAQ #3 (CNV_PE451)\nCNV_PE_SUPPORT_PROJ_GEN_TRULES'
    },
    {
        'msg_class': 'CNV_OT_CX',
        'msg_no': '000',
        'count': '1746',
        'text': 'Generic exception message (&V1 &V2 &V3 &V4)',
        'severity': 'HIGH',
        'root_cause': 'This is a generic exception catch-all message class. The actual error text is in the variable fields (&V1-&V4). This typically indicates an unhandled exception during processing - could be ABAP dump, RFC error, authorization issue, or data inconsistency in the target system.',
        'solution': '1. Click on individual error entries in the Error Monitor to see the actual error text (the &V1-&V4 variables will be filled with specifics).\n2. Check ST22 (ABAP dumps) in both the Control System and Target System for related dumps.\n3. Check SM21 (system log) for additional error details.\n4. If RFC-related: verify RFC destinations in SM59 and check RFC user authorizations.\n5. After identifying root cause, fix and re-generate if needed, then rebuild and re-process.',
        'reference': 'Check ST22 dumps in target\nSM21 system log'
    },
    {
        'msg_class': 'F5A',
        'msg_no': '002',
        'count': '86',
        'text': 'Vendor account is flagged for deletion',
        'severity': 'MEDIUM',
        'root_cause': 'The vendor master record in the target system (MWA) has a deletion flag set. SAP does not allow posting to accounts marked for deletion. This is a master data issue in the target system.',
        'solution': '1. Identify the affected vendor accounts from the error details.\n2. In target system, use XK02/FK02 to remove the deletion flag from the vendor master (General Data -> Status tab or Company Code data).\n3. Alternatively, if deletion is intentional, add a Skip Rule or WL1 Modification Rule to exclude these vendors from migration.\n4. Re-process affected worklist items after master data correction.',
        'reference': 'XK02/FK02 -> Remove deletion flag\nOr implement Skip Rule'
    },
    {
        'msg_class': 'CNV_PE',
        'msg_no': '589',
        'count': '54',
        'text': 'Error executing transfer list modification rule',
        'severity': 'MEDIUM',
        'root_cause': 'The WL2 Transfer List Modification Rule (exit class) encountered an error during execution. This is custom ABAP code that modifies the transfer list before posting. The error could be in the ABAP logic itself, missing data, or an incorrect class type assignment.',
        'solution': '1. Debug the WL2MOD exit class: set breakpoint in the modification rule class and trace through with a failing worklist item.\n2. Ensure the exit class inherits from /SLO/CL_PECON_GEN_MAP (see FAQ #12).\n3. Check that all required data is available in the transfer list structure.\n4. Review the buffer logic if using general buffer pattern for exit classes.\n5. After fixing the ABAP code, regenerate and rebuild transfer list.',
        'reference': 'PECON FAQ #12\nCNV_PE_PROJ -> Transfer Method -> Exits'
    },
    {
        'msg_class': 'CNV_PE',
        'msg_no': '205',
        'count': '51',
        'text': 'Exception in processing / Reference document not yet posted',
        'severity': 'MEDIUM',
        'root_cause': 'The worklist item references another document (via REBZG field) that has not yet been posted in the target system. This is a dependency issue - the reference document must be posted first before the dependent document can be created.',
        'solution': '1. Identify which reference documents (REBZG) are required by checking error details.\n2. Ensure reference documents are processed and posted BEFORE dependent items. Use the GROUP_ID configuration or Plan Jobs to control processing order.\n3. If the reference document was already posted, run "Result Link Update" to propagate the new document number, then rebuild and re-process.\n4. Consider using Predecessor Areas if the reference documents are in a different migration object.',
        'reference': 'PECON FAQ #6 (CNV_PE205)\nGROUP_ID config parameter'
    },
    {
        'msg_class': 'F5',
        'msg_no': '351',
        'count': '17',
        'text': 'Account is blocked for posting',
        'severity': 'MEDIUM',
        'root_cause': 'The G/L account or vendor/customer account in the target system is blocked for posting. This can be set at company code level in the master data.',
        'solution': '1. Identify the blocked accounts from the error details.\n2. In target system, check the account master:\n   - For G/L: FS00 -> Company Code Data -> check "Blocked for posting" flag\n   - For Vendors: FK02/XK02 -> Company Code data -> Accounting info\n3. Remove the posting block if appropriate, or update the Account List to map to a different (unblocked) account.\n4. Re-process affected worklist items.',
        'reference': 'FS00 / FK02 / XK02\nAccount List mapping'
    },
    {
        'msg_class': 'F5A',
        'msg_no': '003',
        'count': '14',
        'text': 'G/L account is flagged for deletion',
        'severity': 'MEDIUM',
        'root_cause': 'The G/L account in the target system has a deletion flag. SAP blocks postings to accounts marked for deletion.',
        'solution': '1. Identify affected G/L accounts from the error details.\n2. In target system, use FS00 to check and remove the deletion flag from the G/L account master.\n3. Alternatively, update the Account List to map source accounts to different target G/L accounts that are active.\n4. Re-process affected worklist items after correction.',
        'reference': 'FS00 -> Remove deletion flag\nAccount List remapping'
    },
    {
        'msg_class': 'F5',
        'msg_no': '026',
        'count': '10',
        'text': 'Vendor has no bank details with bank type',
        'severity': 'LOW',
        'root_cause': 'The vendor master in the target system is missing bank details for the expected bank type. This occurs when open items include payment-relevant data that require bank details.',
        'solution': '1. Identify the affected vendors and required bank types from error details.\n2. In target system, use FK02/XK02 -> Payment Transactions tab to maintain the required bank details for the vendor.\n3. If bank details migration is handled separately, ensure it completes before running PE transfer.\n4. Re-process affected worklist items after master data update.',
        'reference': 'FK02/XK02 -> Payment Transactions\nBank master data'
    },
    {
        'msg_class': '/SLO/PECON',
        'msg_no': '102',
        'count': '5',
        'text': 'Vendor master withholding tax data not maintained',
        'severity': 'LOW',
        'root_cause': 'The vendor master record in the target system does not have withholding tax (WHT) data maintained, but the open item being migrated contains withholding tax information.',
        'solution': '1. Identify affected vendors from the error details.\n2. In target system, use FK02/XK02 -> Withholding Tax tab to maintain the required withholding tax types and codes.\n3. Verify that WHT types in the Account List / transformation match what is configured in target system customizing (OBWW/OBWI).\n4. See also "Special Topic Withholding Tax Handling" section in the PECON User Guide for ACDOCGEN_FI_AP_OI_RCV.\n5. Re-process after WHT master data is corrected.',
        'reference': 'PECON User Guide 7.1.2.1.2\nFK02 -> Withholding Tax tab\nOBWW/OBWI customizing'
    },
]

# Write data rows
for idx, err in enumerate(errors, 1):
    row = idx + 4
    ws[f'A{row}'] = idx
    ws[f'B{row}'] = err['msg_class']
    ws[f'C{row}'] = err['msg_no']
    ws[f'D{row}'] = err['count']
    ws[f'E{row}'] = err['text']
    ws[f'F{row}'] = err['severity']
    ws[f'G{row}'] = err['root_cause']
    ws[f'H{row}'] = err['solution']
    ws[f'I{row}'] = err['reference']

    for col in 'ABCDEFGHI':
        cell = ws[f'{col}{row}']
        cell.alignment = wrap_align
        cell.border = thin_border
        cell.font = Font(name='Calibri', size=10)

    severity_cell = ws[f'F{row}']
    if err['severity'] == 'HIGH':
        severity_cell.fill = severity_high_fill
        severity_cell.font = Font(name='Calibri', size=10, bold=True, color='9C0006')
    elif err['severity'] == 'MEDIUM':
        severity_cell.fill = severity_med_fill
        severity_cell.font = Font(name='Calibri', size=10, bold=True, color='9C6500')
    else:
        severity_cell.fill = severity_low_fill
        severity_cell.font = Font(name='Calibri', size=10, bold=True, color='006100')

    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='top')
    ws[f'D{row}'].alignment = Alignment(horizontal='right', vertical='top')
    ws[f'F{row}'].alignment = Alignment(horizontal='center', vertical='top')

for row in range(5, 15):
    ws.row_dimensions[row].height = 100
ws.row_dimensions[4].height = 30

# --- Summary sheet ---
ws2 = wb.create_sheet('Summary')

ws2.merge_cells('A1:D1')
ws2['A1'] = 'Error Summary - ZFI_SDT / ZAP_OI (Vendor Open Items)'
ws2['A1'].font = title_font

headers2 = [('A', 'Severity', 12), ('B', 'Error Count', 15), ('C', 'Affected Items', 18), ('D', 'Action Required', 60)]
for col_letter, header_text, width in headers2:
    cell = ws2[f'{col_letter}3']
    cell.value = header_text
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border
    ws2.column_dimensions[col_letter].width = width

summary_data = [
    ('HIGH', 3, '9999+ / 2328 / 1746', 'Fix number ranges in target, investigate TRule failures and generic exceptions. These 3 errors account for the vast majority of failures.'),
    ('MEDIUM', 5, '86 / 54 / 51 / 17 / 14', 'Master data corrections (deletion flags, posting blocks), fix WL2MOD exit, resolve document dependencies.'),
    ('LOW', 2, '10 / 5', 'Maintain vendor bank details and withholding tax data in target master records.'),
]

for idx, (sev, cnt, items, action) in enumerate(summary_data, 4):
    ws2[f'A{idx}'] = sev
    ws2[f'B{idx}'] = cnt
    ws2[f'C{idx}'] = items
    ws2[f'D{idx}'] = action
    for col in 'ABCD':
        cell = ws2[f'{col}{idx}']
        cell.alignment = wrap_align
        cell.border = thin_border
        cell.font = Font(name='Calibri', size=10)

    if sev == 'HIGH':
        ws2[f'A{idx}'].fill = severity_high_fill
    elif sev == 'MEDIUM':
        ws2[f'A{idx}'].fill = severity_med_fill
    else:
        ws2[f'A{idx}'].fill = severity_low_fill

ws2.row_dimensions[4].height = 40
ws2.row_dimensions[5].height = 40
ws2.row_dimensions[6].height = 30

# Recommended priority order
ws2.merge_cells('A8:C8')
ws2['A8'] = 'Recommended Resolution Order'
ws2['A8'].font = Font(name='Calibri', size=12, bold=True, color='1F4E79')

priority_headers = [('A', 'Priority', 15), ('B', 'Error', 40), ('C', 'Action', 60)]
for col_letter, header_text, width in priority_headers:
    cell = ws2[f'{col_letter}9']
    cell.value = header_text
    cell.font = header_font
    cell.fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    cell.alignment = header_align
    cell.border = thin_border

priority = [
    ('1 (Critical)', 'NR 751 - Number range intervals', 'Blocks 9999+ items. Quick fix in target customizing (FBN1/SNRO).'),
    ('2 (Critical)', 'CNV_PE 451 - TRule errors', 'Blocks 2328 items. Read Customizing + regenerate TRules.'),
    ('3 (Investigate)', 'CNV_OT_CX 000 - Generic exceptions', 'Blocks 1746 items. Check ST22/SM21 for actual root cause.'),
    ('4 (Master Data)', 'F5A 002/003, F5 351 - Deletion flags & blocks', 'Fix in target vendor/GL masters (XK02, FS00).'),
    ('5 (Dependencies)', 'CNV_PE 205 - Reference doc not posted', 'Control processing order, post reference docs first.'),
    ('6 (Code Fix)', 'CNV_PE 589 - WL2MOD exit error', 'Debug and fix custom ABAP exit class.'),
    ('7 (Master Data)', 'F5 026, /SLO/PECON 102 - Bank & WHT data', 'Maintain vendor bank details and WHT data.'),
]

for idx, (prio, error, action) in enumerate(priority, 10):
    ws2[f'A{idx}'] = prio
    ws2[f'B{idx}'] = error
    ws2[f'C{idx}'] = action
    for col in 'ABC':
        cell = ws2[f'{col}{idx}']
        cell.alignment = wrap_align
        cell.border = thin_border
        cell.font = Font(name='Calibri', size=10)
    ws2.row_dimensions[idx].height = 30

# Freeze panes
ws.freeze_panes = 'A5'
ws2.freeze_panes = 'A3'

# Auto-filter
ws.auto_filter.ref = f'A4:I{4 + len(errors)}'

# Save
filepath = r'C:\Users\I557430\Documents\sap_gui_scripting_ai\PE_Error_Analysis_ZFI_SDT.xlsx'
wb.save(filepath)
print(f'Excel file saved to: {filepath}')
print(f'Sheets: Error Analysis ({len(errors)} errors), Summary (with priority order)')
